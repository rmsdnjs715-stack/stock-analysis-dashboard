#!/usr/bin/env python3
"""주식분석 자동화 - 메인 실행 스크립트.

시트1(매크로지표)·시트2(빅테크Top20)를 Yahoo Finance/FRED 최신 데이터로 갱신한다.
시트3(뉴스정리)·시트4(주식격언)는 WebSearch 등 Claude 전용 도구가 필요해 이 스크립트
범위 밖이다 - 기존 파일에 있으면 그대로 보존한다.

사용법:
    python main.py                 # output/주식시장_분석리포트.xlsx 갱신
    python main.py --add-ticker PLTR   # 커스텀 종목 추가 후 갱신
"""
from __future__ import annotations

import argparse
import sys
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent / "src"))

from openpyxl import Workbook, load_workbook  # noqa: E402

import config  # noqa: E402
from build_excel import build_bigtech_sheet, build_macro_sheet  # noqa: E402
from fetch_equities import fetch_top_bigtech  # noqa: E402
from fetch_macro import fetch_all_macro  # noqa: E402

# 이 프로젝트에서 동시에 진행 중인 다른 자동화가 config.REPORT_FILE(공유 기본 경로)에
# 별도 구조의 시트를 계속 쓰고 있어, 이 리포트(매크로+빅테크Top20+뉴스+격언, 차트 포함)는
# 충돌을 피하기 위해 고유 파일명을 따로 쓴다.
REPORT_FILE = config.OUTPUT_DIR / "빅테크_매크로_분석리포트.xlsx"

SHEET1_NAME = "1.매크로지표"
SHEET2_NAME = "2.빅테크Top20"
SHEET3_NAME = "3.뉴스정리"
SHEET4_NAME = "4.주식격언"
SHEET_ORDER = [SHEET1_NAME, SHEET2_NAME, SHEET3_NAME, SHEET4_NAME]


def _load_or_create_workbook() -> Workbook:
    if REPORT_FILE.exists():
        return load_workbook(REPORT_FILE)
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def _reset_sheet(wb: Workbook, name: str):
    if name in wb.sheetnames:
        del wb[name]
    idx = SHEET_ORDER.index(name)
    return wb.create_sheet(name, index=min(idx, len(wb.sheetnames)))


def main() -> None:
    parser = argparse.ArgumentParser(description="빅테크/매크로 데이터로 리포트 시트1·2 갱신")
    parser.add_argument("--add-ticker", help="빅테크 시트에 항상 포함시킬 커스텀 종목 티커 (예: PLTR)")
    args = parser.parse_args()

    if args.add_ticker:
        config.add_custom_ticker(args.add_ticker)
        print(f"커스텀 종목 추가: {args.add_ticker.upper()} -> {config.CUSTOM_TICKERS_PATH}")

    today = date.today()

    print("[1/4] 매크로 지표 수집 중...")
    macro_data = fetch_all_macro()
    ok = sum(1 for d in macro_data if not d["error"])
    print(f"   -> {ok}/{len(macro_data)}개 성공")

    print("[2/4] 빅테크 Top20 데이터 수집 중 (수 분 소요될 수 있음)...")
    equities = fetch_top_bigtech()
    print(f"   -> {len(equities)}종목 수집 완료")

    print("[3/4] 엑셀 시트 작성 중...")
    config.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = _load_or_create_workbook()

    ws1 = _reset_sheet(wb, SHEET1_NAME)
    build_macro_sheet(ws1, macro_data, today)

    ws2 = _reset_sheet(wb, SHEET2_NAME)
    build_bigtech_sheet(ws2, equities, today)

    # 시트3/4가 아직 없다면 자리만 만들어둔다 (Claude가 뉴스/격언 리서치 후 채움)
    for name in (SHEET3_NAME, SHEET4_NAME):
        if name not in wb.sheetnames:
            ws = wb.create_sheet(name)
            ws["A1"] = f"{name} - 아직 데이터가 채워지지 않았습니다."

    print(f"[4/4] 저장: {REPORT_FILE}")
    wb.save(REPORT_FILE)
    print("완료.")


if __name__ == "__main__":
    main()
