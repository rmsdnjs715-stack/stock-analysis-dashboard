"""시트3(뉴스)·시트4(격언) 작성 - WebSearch 리서치 결과(news_items)와 격언 파일 병합
결과(quote_rows)를 받아 기존 리포트 워크북에 반영한다.
이 스크립트는 Claude가 WebSearch로 수집한 데이터를 넘겨받아 실행하는 용도이며,
main.py(순수 API 자동화)와 분리되어 있다.
"""
from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from openpyxl import load_workbook

import config
from build_excel import build_news_sheet, build_quotes_sheet

SHEET3_NAME = "3.뉴스정리"
SHEET4_NAME = "4.주식격언"

# main.py와 동일한 고유 파일명 사용 (공유 config.REPORT_FILE 경로는 이 프로젝트에서
# 동시에 진행 중인 다른 자동화가 쓰고 있어 충돌 방지 차원에서 분리함)
REPORT_FILE = config.OUTPUT_DIR / "빅테크_매크로_분석리포트.xlsx"

NEWS_ITEMS: list[dict] = [
    {
        "source": "Bloomberg / CNBC",
        "headline": "Big Tech Stocks Storm Back as AI Fears Fade and Euphoria Resumes",
        "summary": "MSFT·AMZN의 AI 매출 성장 가속 실적을 계기로 빅테크 랠리 재개. "
                   "다우 53,178pt·S&P500 7,600pt대 사상 최고치 경신, 나스닥 +2.1%.",
        "related": "MSFT, AMZN, 나스닥",
        "url": "https://www.bloomberg.com/news/articles/2026-08-07/big-tech-stocks-storm-back-as-ai-fears-fade-and-euphoria-resumes",
    },
    {
        "source": "Fortune (WSJ 등 종합)",
        "headline": "Big Tech's AI Capex to Reach ~$725~760B in 2026 (전년比 +77%)",
        "summary": "MSFT(~$190B)·AMZN(~$200B)·GOOGL($175~205B)·META($115~135B) 합산 "
                   "AI 인프라 투자 급증. 2027년 $1조 돌파 전망. 클라우드 백로그 $2.3조로 확대.",
        "related": "MSFT, AMZN, GOOGL, META",
        "url": "https://fortune.com/2026/07/26/big-tech-earnings-meta-microsoft-apple-amazon-market-revolt-ai-spending/",
    },
    {
        "source": "Fortune / Yahoo Finance",
        "headline": "Alphabet EPS $9.11 어닝서프라이즈에도 주가 -7% (자본지출 확대·FCF 첫 마이너스)",
        "summary": "2분기 EPS가 컨센서스($3.00)를 크게 상회했지만, 2026년 capex를 최대 $205B로 "
                   "상향하고 FCF가 2004년 상장 후 최초로 마이너스 전환하며 주가 급락.",
        "related": "GOOGL",
        "url": "https://fortune.com/2026/07/26/big-tech-earnings-meta-microsoft-apple-amazon-market-revolt-ai-spending/",
    },
    {
        "source": "CNBC",
        "headline": "Microsoft FY26 4분기 실적 컨센서스 상회 - Azure·AI 성장 견인",
        "summary": "Azure 클라우드와 AI 관련 매출 호조로 시장 예상치를 상회하는 실적 발표.",
        "related": "MSFT",
        "url": "https://www.cnbc.com/2026/01/27/big-tech-earnings-2026-ai-spend.html",
    },
    {
        "source": "Fortune",
        "headline": "Meta, 실적 부진 + 시장은 완만한 capex 상향 가이던스를 신뢰하지 않음",
        "summary": "예상보다 낮은 영업이익률과 소폭의 2026 capex 상향 발표에도 시장 반응은 부정적.",
        "related": "META",
        "url": "https://fortune.com/2026/07/26/big-tech-earnings-meta-microsoft-apple-amazon-market-revolt-ai-spending/",
    },
    {
        "source": "Seeking Alpha",
        "headline": "Apple's Selloff Looks More Like An Opportunity Than A Warning",
        "summary": "실적 후 조정은 수요 둔화가 아닌 공급 병목 때문이라며 Buy 의견, 목표가 $350 "
                   "(내재가치 대비 +13% 여력). 6월 분기 매출 +16%, 아이폰 판매 +22%.",
        "related": "AAPL",
        "url": "https://seekingalpha.com/article/4929301-apples-selloff-looks-more-like-an-opportunity-than-a-warning",
    },
    {
        "source": "Seeking Alpha",
        "headline": "Nvidia: Upside Is Starting To Fade",
        "summary": "AI/데이터센터 강세와 대규모 투자 파이프라인에도 밸류에이션 부담을 지적하며 "
                   "Hold 의견, 목표가 $244 제시.",
        "related": "NVDA",
        "url": "https://seekingalpha.com/article/4932374-nvidia-upside-is-starting-to-fade",
    },
    {
        "source": "Yahoo Finance / MarketBeat (내부자 거래)",
        "headline": "빅테크 대주주·임원 매매 동향: 뚜렷한 '대량 매수' 신호는 부재",
        "summary": "Palantir(PLTR)는 최근 90일간 이사(Alexander Moore) 1.6만주 매도 등 매도 우위. "
                   "Oracle(ORCL)은 RSU 지급($60,184) 등 통상적 활동만 포착, 유의미한 매수 신호 없음.",
        "related": "PLTR, ORCL",
        "url": "https://finance.yahoo.com/quote/PLTR/insider-transactions/",
    },
    {
        "source": "24/7 Wall St. / Yahoo Finance",
        "headline": "AI 반도체 슈퍼사이클: Broadcom·Micron·TSMC 실적 가이던스 급등",
        "summary": "Broadcom AI반도체 매출 가이던스 +200%YoY($16.0B), Micron 4분기 매출 가이던스 "
                   "$50.0B(HBM4 본격 출하), TSMC 2026 매출 +30%YoY 전망. 5월 전세계 반도체 "
                   "매출 $120.6B(+104%YoY) 기록.",
        "related": "AVGO, MU, TSM",
        "url": "https://247wallst.com/investing/2026/08/01/3-semiconductor-stocks-to-buy-before-ai-demand-explodes-in-august/",
    },
    {
        "source": "iShares / Yahoo Finance (매크로)",
        "headline": "Fed, 2026년 금리인하 無 - 신임 의장 매파적 스탠스로 변동성 확대 우려",
        "summary": "기준금리 3.50~3.75% 유지 중 연내 인하 없이 동결 기조. 신임 Fed 의장 Kevin Warsh는 "
                   "포워드가이던스 최소화를 선호해 금리 변동성 확대 가능성. 일부는 오히려 "
                   "인상 가능성까지 가격에 반영.",
        "related": "전체 매크로 / VIX / 금리",
        "url": "https://www.ishares.com/us/insights/portfolio-insights/fed-outlook-rates-kevin-warsh-fixed-income-2026",
    },
]

# 종목별 "최근 이슈" 한 줄 요약 (시트2 마지막 열에 반영)
RECENT_INSIGHTS: dict[str, str] = {
    "AAPL": "실적 후 조정은 수요둔화 아닌 공급병목 - SA는 Buy/PT $350 제시(2026-08)",
    "GOOGL": "2Q EPS $9.11 서프라이즈에도 capex 최대 $205B 확대+FCF 첫 마이너스로 주가 -7%",
    "MSFT": "FY26 4Q Azure·AI 매출 호조로 컨센서스 상회, capex ~$190B",
    "AMZN": "AI 매출 성장 가속 실적이 빅테크 랠리 재점화 트리거 중 하나(2026-08)",
    "META": "2Q 실적 컨센서스 하회, capex 소폭 상향에도 시장 신뢰 회복 못함",
    "NVDA": "SA는 Hold/PT $244 - AI 강세는 유효하나 밸류에이션 부담 지적",
    "AVGO": "AI반도체 매출 가이던스 +200%YoY($16.0B) - 커스텀 가속기 포지셔닝 부각",
    "TSM": "2026 매출 +30%YoY 전망, AI 가속기 수요 견조",
    "MU": "4Q 매출가이던스 $50.0B, HBM4 본격 출하로 메모리 슈퍼사이클 수혜",
    "ORCL": "8/3 기준 RSU 지급 외 유의미한 대주주 매수 신호 없음(통상적 활동)",
    "PLTR": "최근 90일 이사 1.6만주 매도 등 내부자 매도 우위, 뚜렷한 매수 신호 부재",
    "TSLA": "Palantir와 함께 월가 '$13B 경고' 헤드라인 등장 - 밸류에이션 주의 필요",
}


def _load_quote_rows_with_buffett() -> list[tuple[str, str, str]]:
    """기존 매매격언 시트를 읽고, 버핏 명언을 새 구분('버핏')으로 append.
    이미 '버핏' 구분이 존재하면 재실행해도 중복 추가하지 않는다(idempotent)."""
    wb = load_workbook(config.QUOTES_FILE)
    ws = wb["매매격언"]
    rows: list[tuple[str, str, str]] = []
    last_gubun = ""
    already_has_buffett = False
    for r in ws.iter_rows(min_row=2, values_only=True):  # 2행부터 (1행은 타이틀)
        if r[0] is None and r[1] is None:
            continue
        gubun = r[0] if r[0] else last_gubun
        if r[0]:
            last_gubun = r[0]
        if gubun == "버핏":
            already_has_buffett = True
        content = r[1] or ""
        note = r[2] if len(r) > 2 and r[2] else ""
        if content:
            rows.append((gubun, content, note))

    if already_has_buffett:
        print("버핏 명언이 이미 반영되어 있어 다시 추가하지 않음 (idempotent)")
        return rows

    buffett_quotes = [
        "잘 모르는 분야에는 투자하지 않는다.",
        "한 가지 제품에 회사의 운명이 좌우되는 기업에는 투자하지 않는다.",
        "부채가 많은 기업은 쳐다보지 않는다.",
        "매출이 많아도 나쁜 기업은 피한다.",
        "경영진이 정직하고 능력있는 회사를 선택한다.",
        "단기적인 시세차익을 바라고 매수하지 않는다.",
        "일단 자신의 투자방식을 신뢰한다.",
        "좋아하는 공이 올 때까지 기다린다.",
        "시장이 폭락할 때는 바겐세일이 온 것이다.",
        "주가를 거래하지 말고 기업을 거래하라.",
        "주가순유동자산(순유동자산가치) 2/3 이하 종목에 투자하라.",
    ]
    # 새 xlsx에 append
    next_row = ws.max_row + 1
    for q in buffett_quotes:
        ws.cell(row=next_row, column=1, value="버핏")
        ws.cell(row=next_row, column=2, value=q)
        ws.cell(row=next_row, column=3, value="워런 버핏 투자원칙")
        rows.append(("버핏", q, "워런 버핏 투자원칙"))
        next_row += 1
    wb.save(config.QUOTES_FILE)
    return rows


def apply_recent_insights(ws) -> None:
    """시트2(빅테크Top20)의 티커별 '최근이슈' 열(20번째 열)을 채운다."""
    header_row = 4
    ticker_col = 2
    insight_col = 20
    for row in range(header_row + 1, ws.max_row + 1):
        ticker = ws.cell(row=row, column=ticker_col).value
        if ticker in RECENT_INSIGHTS:
            ws.cell(row=row, column=insight_col, value=RECENT_INSIGHTS[ticker])


def main() -> None:
    if not REPORT_FILE.exists():
        raise SystemExit(f"리포트 파일이 없습니다. 먼저 main.py를 실행하세요: {REPORT_FILE}")

    today = date.today()
    wb = load_workbook(REPORT_FILE)

    if SHEET3_NAME in wb.sheetnames:
        del wb[SHEET3_NAME]
    ws3 = wb.create_sheet(SHEET3_NAME, index=2)
    build_news_sheet(ws3, NEWS_ITEMS, today)

    print("매매격언 원본 파일에 버핏 명언 추가 중...")
    quote_rows = _load_quote_rows_with_buffett()

    if SHEET4_NAME in wb.sheetnames:
        del wb[SHEET4_NAME]
    ws4 = wb.create_sheet(SHEET4_NAME, index=3)
    build_quotes_sheet(ws4, quote_rows)

    if "2.빅테크Top20" in wb.sheetnames:
        apply_recent_insights(wb["2.빅테크Top20"])

    wb.save(REPORT_FILE)
    print(f"완료: {REPORT_FILE}")


if __name__ == "__main__":
    main()
