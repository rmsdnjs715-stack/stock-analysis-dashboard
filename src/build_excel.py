"""openpyxl로 표 + 네이티브 엑셀 차트를 만드는 빌더 모음.
시트1(매크로), 시트2(빅테크Top20)는 데이터만 넣으면 자동으로 표+차트 생성.
시트3(뉴스), 시트4(격언)는 리서치/병합된 텍스트 데이터를 받아 표로만 정리.
"""
from __future__ import annotations

from datetime import date
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=16)
SUBTLE_FONT = Font(italic=True, size=9, color="6B7280")


def _style_header_row(ws: Worksheet, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autosize(ws: Worksheet, widths: dict[int, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


# ---------------------------------------------------------------- 시트1: 매크로 ----
def build_macro_sheet(ws: Worksheet, macro_data: list[dict[str, Any]], as_of: date) -> None:
    ws["A1"] = "매크로 지표 대시보드"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"기준일: {as_of.isoformat()}  (자동수집 실패 항목은 웹 검색으로 보완 표기)"
    ws["A2"].font = SUBTLE_FONT

    header_row = 4
    headers = ["지표", "최신값", "단위", "기준일", "비고"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=i, value=h)
    _style_header_row(ws, header_row, len(headers))

    row = header_row + 1
    for ind in macro_data:
        ws.cell(row=row, column=1, value=ind["name"])
        val = ind.get("latest_value")
        ws.cell(row=row, column=2, value=round(val, 2) if isinstance(val, (int, float)) else "N/A")
        ws.cell(row=row, column=3, value=ind["unit"])
        d = ind.get("latest_date")
        ws.cell(row=row, column=4, value=d.isoformat() if d else "-")
        note = "자동 수집 실패 - 수동 확인 필요" if ind.get("error") else (ind.get("note") or "")
        ws.cell(row=row, column=5, value=note)
        row += 1

    _autosize(ws, {1: 26, 2: 14, 3: 12, 4: 14, 5: 34})
    ws.freeze_panes = "A5"

    # ---- 원본 히스토리 데이터 (차트 소스, 시각적으로 멀리 배치) ----
    data_start_col = 8  # H열부터
    chart_anchor_row = header_row + 2
    chart_col_step = 9
    chart_row_step = 10
    charts_per_row = 2

    col = data_start_col
    for idx, ind in enumerate(macro_data):
        # 차트/원본표는 최근 구간만 (파일 용량 절감 - 최신값·비교는 이미 요약표에 있음)
        hist = (ind.get("history") or [])[-104:]
        ws.cell(row=1, column=col, value=ind["name"]).font = Font(bold=True)
        ws.cell(row=2, column=col, value="날짜")
        ws.cell(row=2, column=col + 1, value="값")
        for i, (d, v) in enumerate(hist, start=3):
            ws.cell(row=i, column=col, value=d.isoformat())
            ws.cell(row=i, column=col + 1, value=v)

        if hist:
            chart = LineChart()
            chart.title = ind["name"]
            chart.height, chart.width = 7, 11
            chart.style = 2
            data_ref = Reference(ws, min_col=col + 1, min_row=2, max_row=2 + len(hist))
            cats_ref = Reference(ws, min_col=col, min_row=3, max_row=2 + len(hist))
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.y_axis.title = ind["unit"]
            chart.legend = None

            grid_pos = idx % charts_per_row
            grid_row = idx // charts_per_row
            anchor_col_letter = get_column_letter(1 + grid_pos * chart_col_step)
            anchor_row = chart_anchor_row + grid_row * chart_row_step
            ws.add_chart(chart, f"{anchor_col_letter}{anchor_row}")

        col += 3  # 데이터 2열 + 여백 1열


# ------------------------------------------------------------ 시트2: 빅테크 Top20 ----
_BIGTECH_HEADERS = [
    "순위", "티커", "기업명", "시가총액($B)", "현재가($)", "EPS", "ROE(%)",
    "현금성자산($B)", "PER", "전고점대비(%)", "MA20", "MA60", "MA120", "MA200",
    "이평배열", "RSI(14)", "RSI상태", "MACD", "MACD상태", "최근이슈",
]


def build_bigtech_sheet(
    ws: Worksheet, equities: list[dict[str, Any]], as_of: date, chart_top_n: int = 20
) -> None:
    ws["A1"] = "빅테크 시가총액 Top20 (+커스텀 추가 종목)"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"기준일: {as_of.isoformat()}  (기술적지표: 최근 400거래일 종가 기준, 전고점: 상장 이후 전체 기간)"
    ws["A2"].font = SUBTLE_FONT

    header_row = 4
    for i, h in enumerate(_BIGTECH_HEADERS, start=1):
        ws.cell(row=header_row, column=i, value=h)
    _style_header_row(ws, header_row, len(_BIGTECH_HEADERS))

    def fmt(v, nd=2):
        return round(v, nd) if isinstance(v, (int, float)) else "N/A"

    row = header_row + 1
    for rank, eq in enumerate(equities, start=1):
        t = eq.get("technical", {}) or {}
        vals = [
            rank,
            eq["ticker"],
            eq["name"],
            fmt(eq["market_cap"] / 1e9 if eq.get("market_cap") else None, 1),
            fmt(eq.get("current_price")),
            fmt(eq.get("eps")),
            fmt(eq.get("roe_pct"), 1),
            fmt(eq["cash"] / 1e9 if eq.get("cash") else None, 1),
            fmt(eq.get("per"), 1),
            fmt(eq.get("drawdown_pct"), 1),
            fmt(t.get("ma20")),
            fmt(t.get("ma60")),
            fmt(t.get("ma120")),
            fmt(t.get("ma200")),
            t.get("cross_20_60", "N/A"),
            fmt(t.get("rsi14"), 1),
            t.get("rsi_regime", "N/A"),
            fmt(t.get("macd")),
            t.get("macd_regime", "N/A"),
            eq.get("recent_insight", ""),
        ]
        for c, v in enumerate(vals, start=1):
            ws.cell(row=row, column=c, value=v)
        row += 1

    widths = {1: 6, 2: 10, 3: 26, 4: 14, 5: 12, 6: 9, 7: 9, 8: 14, 9: 9, 10: 12,
              11: 10, 12: 10, 13: 10, 14: 10, 15: 10, 16: 10, 17: 9, 18: 10, 19: 10, 20: 40}
    _autosize(ws, widths)
    ws.freeze_panes = "A5"

    # ---- 종목별 주가 히스토리(월말 종가) + 미니 차트 ----
    # 파일 용량 때문에 개별 차트+원본데이터는 시총 상위 chart_top_n개만 (나머지는 위 표에
    # 수치는 다 있음, 차트만 생략). 요청 시 chart_top_n을 늘리면 더 많은 종목도 가능.
    charted = equities[:chart_top_n]
    if len(charted) < len(equities):
        ws.cell(row=3, column=1, value=f"※ 개별 주가 차트는 시총 상위 {chart_top_n}종목만 표시 (전 종목 수치는 위 표 참고)")
        ws.cell(row=3, column=1).font = SUBTLE_FONT

    data_start_col = len(_BIGTECH_HEADERS) + 3
    chart_anchor_row = header_row + 2
    chart_col_step = 9
    chart_row_step = 10
    charts_per_row = 3

    col = data_start_col
    for idx, eq in enumerate(charted):
        # 차트/원본표는 최근 24개월만 (전고점/하락률은 이미 표에 반영되어 있어
        # 상장 이후 전체 히스토리를 여기 다 실을 필요는 없음 - 파일 용량 절감)
        hist = (eq.get("price_history") or [])[-24:]
        label = f"{eq['ticker']} 주가"
        ws.cell(row=1, column=col, value=label).font = Font(bold=True)
        ws.cell(row=2, column=col, value="월")
        ws.cell(row=2, column=col + 1, value="종가")
        for i, (d, v) in enumerate(hist, start=3):
            ws.cell(row=i, column=col, value=d.strftime("%Y-%m"))
            ws.cell(row=i, column=col + 1, value=v)

        if hist:
            chart = LineChart()
            chart.title = label
            chart.height, chart.width = 6.5, 9
            chart.style = 10
            data_ref = Reference(ws, min_col=col + 1, min_row=2, max_row=2 + len(hist))
            cats_ref = Reference(ws, min_col=col, min_row=3, max_row=2 + len(hist))
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.legend = None

            grid_pos = idx % charts_per_row
            grid_row = idx // charts_per_row
            anchor_col_letter = get_column_letter(1 + grid_pos * chart_col_step)
            anchor_row = chart_anchor_row + grid_row * chart_row_step
            ws.add_chart(chart, f"{anchor_col_letter}{anchor_row}")

        col += 3


# --------------------------------------------------------------------- 시트3: 뉴스 ----
_NEWS_HEADERS = ["출처", "제목/헤드라인", "요약", "관련 빅테크", "링크", "수집일"]


def build_news_sheet(ws: Worksheet, news_items: list[dict[str, Any]], as_of: date) -> None:
    ws["A1"] = "뉴스 정리 (월가 · 파이낸셜타임스 · Seeking Alpha)"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"수집일: {as_of.isoformat()}  (유료 구독 사이트 특성상 헤드라인/요약 수준으로 정리, 전문은 원문 링크 참고)"
    ws["A2"].font = SUBTLE_FONT

    header_row = 4
    for i, h in enumerate(_NEWS_HEADERS, start=1):
        ws.cell(row=header_row, column=i, value=h)
    _style_header_row(ws, header_row, len(_NEWS_HEADERS))

    row = header_row + 1
    for item in news_items:
        ws.cell(row=row, column=1, value=item.get("source", ""))
        ws.cell(row=row, column=2, value=item.get("headline", ""))
        ws.cell(row=row, column=3, value=item.get("summary", ""))
        ws.cell(row=row, column=4, value=item.get("related", ""))
        ws.cell(row=row, column=5, value=item.get("url", ""))
        ws.cell(row=row, column=6, value=as_of.isoformat())
        for c in range(1, 7):
            ws.cell(row=row, column=c).alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    _autosize(ws, {1: 16, 2: 40, 3: 55, 4: 20, 5: 40, 6: 12})
    ws.freeze_panes = "A5"


# --------------------------------------------------------------------- 시트4: 격언 ----
_QUOTE_HEADERS = ["구분", "격언 / 내용", "비고(출처)"]


def build_quotes_sheet(ws: Worksheet, quote_rows: list[tuple[str, str, str]]) -> None:
    ws["A1"] = "주식 매매격언 정리 (원본: 주식매매격언_정리.xlsx > 매매격언 시트 + 버핏 명언 추가)"
    ws["A1"].font = TITLE_FONT

    header_row = 3
    for i, h in enumerate(_QUOTE_HEADERS, start=1):
        ws.cell(row=header_row, column=i, value=h)
    _style_header_row(ws, header_row, len(_QUOTE_HEADERS))

    row = header_row + 1
    for gubun, content, note in quote_rows:
        ws.cell(row=row, column=1, value=gubun)
        ws.cell(row=row, column=2, value=content)
        ws.cell(row=row, column=3, value=note)
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    _autosize(ws, {1: 12, 2: 90, 3: 24})
    ws.freeze_panes = "A4"


def new_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)  # 기본 빈 시트 제거, 아래서 순서대로 새로 추가
    return wb
